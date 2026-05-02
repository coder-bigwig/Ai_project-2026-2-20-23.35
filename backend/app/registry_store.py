from fastapi import HTTPException, UploadFile, File, Request
from fastapi.responses import FileResponse, StreamingResponse, PlainTextResponse
from fastapi.encoders import jsonable_encoder
from pydantic import BaseModel, Field
from typing import Callable, List, Optional, Dict, Tuple, Set
from datetime import datetime, timezone
from enum import Enum
from copy import deepcopy
import csv
import hashlib
import io
import json
import mimetypes
import re
import shutil
import uuid
import os
import zipfile

from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker

from .config import *
from .state import *
from .db.models import (
    AppKVStoreORM,
    AttachmentORM,
    ClassroomORM,
    CourseORM,
    ExperimentORM,
    OperationLogORM,
    ResourceORM,
    SubmissionORM,
    SubmissionPdfORM,
    UserORM,
)
from .integrations.jupyterhub_integration import _hub_user_state_map, _extract_server_state
from .storage_config import DATABASE_URL, use_postgres

_pg_sync_engine = None
_pg_sync_session_maker: Optional[sessionmaker] = None


def _to_sync_driver_url(raw_url: str) -> str:
    url = (raw_url or "").strip()
    if url.startswith("postgresql+psycopg2://"):
        return url
    if url.startswith("postgresql+asyncpg://"):
        return "postgresql+psycopg2://" + url[len("postgresql+asyncpg://") :]
    if url.startswith("postgres://"):
        return "postgresql+psycopg2://" + url[len("postgres://") :]
    return url


def _get_pg_sync_session_maker() -> Optional[sessionmaker]:
    global _pg_sync_engine, _pg_sync_session_maker
    if _pg_sync_session_maker is not None:
        return _pg_sync_session_maker
    if not use_postgres():
        return None
    sync_url = _to_sync_driver_url(DATABASE_URL)
    if not sync_url:
        return None
    _pg_sync_engine = create_engine(sync_url, pool_pre_ping=True, future=True)
    _pg_sync_session_maker = sessionmaker(bind=_pg_sync_engine, autoflush=False, autocommit=False)
    return _pg_sync_session_maker


def _run_pg_write(op_name: str, handler: Callable[[Session], None]) -> None:
    session_maker = _get_pg_sync_session_maker()
    if session_maker is None:
        return
    try:
        with session_maker() as db:
            handler(db)
            db.commit()
    except Exception as exc:
        print(f"[storage] failed to persist {op_name} to postgres: {exc}")


def _upsert_kv(db: Session, key: str, value_json: Optional[dict]) -> None:
    existing = db.get(AppKVStoreORM, key)
    if existing is None:
        db.add(AppKVStoreORM(key=key, value_json=value_json or {}))
        return
    existing.value_json = value_json or {}


def _parse_iso_datetime(value, fallback: Optional[datetime] = None) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        text = value.strip().replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            pass
    return fallback or datetime.now()


def _parse_iso_datetime_nullable(value) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        text = value.strip().replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            return None
    return None

def is_teacher(username: str) -> bool:
    """判断用户是否为教师"""
    normalized = _normalize_text(username)
    if not normalized:
        return False
    if normalized in TEACHER_ACCOUNTS:
        return True

    session_maker = _get_pg_sync_session_maker()
    if session_maker is None:
        return False
    try:
        with session_maker() as db:
            row = (
                db.execute(
                    select(UserORM).where(
                        UserORM.username == normalized,
                        UserORM.role == "teacher",
                    )
                )
                .scalars()
                .first()
            )
            return row is not None
    except Exception:
        return False


def is_admin(username: str) -> bool:
    """判断用户是否为管理员"""
    return username in ADMIN_ACCOUNTS
def _read_seed_marker() -> Tuple[int, dict]:
    """Return (version, payload). Version 0 means not seeded.

    Legacy marker content was a plain timestamp string; treat that as version 1.
    """
    if not os.path.exists(SEED_MARKER_FILE):
        return 0, {}

    try:
        with open(SEED_MARKER_FILE, "r", encoding="utf-8") as file_obj:
            raw = (file_obj.read() or "").strip()
    except OSError as exc:
        print(f"Failed to read seed marker: {exc}")
        return 1, {}

    if not raw:
        return 1, {}

    try:
        payload = json.loads(raw)
    except json.JSONDecodeError:
        return 1, {"legacy": True, "raw": raw}

    if isinstance(payload, dict):
        version = payload.get("version", 1)
        try:
            version = int(version)
        except (TypeError, ValueError):
            version = 1
        return version, payload

    return 1, {}


def _write_seed_marker(version: int, payload: Optional[dict] = None):
    data = {"version": int(version), "updated_at": datetime.now().isoformat()}
    if payload:
        data.update(payload)

    tmp_path = f"{SEED_MARKER_FILE}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as file_obj:
        json.dump(data, file_obj, ensure_ascii=False, indent=2)
    os.replace(tmp_path, SEED_MARKER_FILE)
class DifficultyLevel(str, Enum):
    BEGINNER = "初级"
    INTERMEDIATE = "中级"
    ADVANCED = "高级"

class ExperimentStatus(str, Enum):
    NOT_STARTED = "未开始"
    IN_PROGRESS = "进行中"
    SUBMITTED = "已提交"
    GRADED = "已评分"


class PublishScope(str, Enum):
    ALL = "all"
    CLASS = "class"
    STUDENT = "student"

class Experiment(BaseModel):
    id: str = None
    course_id: Optional[str] = None
    course_name: str = "Python程序设计"
    title: str
    description: Optional[str] = None
    difficulty: DifficultyLevel = DifficultyLevel.BEGINNER
    tags: List[str] = []
    notebook_path: Optional[str] = None
    resources: dict = {"cpu": 1.0, "memory": "2G", "storage": "1G"}
    deadline: Optional[datetime] = None
    created_at: datetime = None
    created_by: str
    published: bool = True  # 是否发布给学生
    publish_scope: PublishScope = PublishScope.ALL
    target_class_names: List[str] = Field(default_factory=list)
    target_student_ids: List[str] = Field(default_factory=list)
    
    class Config:
        json_encoders = {
            datetime: lambda v: v.isoformat()
        }

class StudentExperiment(BaseModel):
    id: str = None
    experiment_id: str
    student_id: str
    status: ExperimentStatus = ExperimentStatus.NOT_STARTED
    start_time: Optional[datetime] = None
    submit_time: Optional[datetime] = None
    notebook_content: Optional[str] = None
    score: Optional[float] = None
    ai_feedback: Optional[str] = None
    teacher_comment: Optional[str] = None

class SubmitExperimentRequest(BaseModel):
    notebook_content: str


class LoginRequest(BaseModel):
    username: str
    password: str


class StudentPasswordChangeRequest(BaseModel):
    student_id: str
    old_password: str
    new_password: str


class StudentSecurityQuestionUpdateRequest(BaseModel):
    student_id: str
    security_question: str
    security_answer: str


class TeacherPasswordChangeRequest(BaseModel):
    teacher_username: str
    old_password: str
    new_password: str


class TeacherSecurityQuestionUpdateRequest(BaseModel):
    teacher_username: str
    security_question: str
    security_answer: str


class ForgotPasswordResetRequest(BaseModel):
    username: str
    security_answer: str
    new_password: str


class CourseCreateRequest(BaseModel):
    name: str
    description: Optional[str] = ""
    teacher_username: str


class CourseUpdateRequest(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    teacher_username: str


class ClassCreateRequest(BaseModel):
    name: str
    teacher_username: str


class TeacherCreateRequest(BaseModel):
    admin_username: str
    username: str
    real_name: Optional[str] = ""


class CourseRecord(BaseModel):
    id: str
    name: str
    description: Optional[str] = ""
    created_by: str
    created_at: datetime
    updated_at: datetime


class ClassRecord(BaseModel):
    id: str
    name: str
    created_by: str
    created_at: datetime


class TeacherRecord(BaseModel):
    username: str
    real_name: str = ""
    created_by: str
    created_at: datetime


class StudentRecord(BaseModel):
    student_id: str
    username: str
    real_name: str
    class_name: str
    admission_year: str = ""
    organization: str
    phone: str
    role: str = "student"
    created_by: str = ""
    password_hash: str
    security_question: str = ""
    security_answer_hash: str = ""
    created_at: datetime
    updated_at: datetime


class PDFAnnotation(BaseModel):
    id: str
    teacher_username: str
    content: str
    created_at: datetime


class StudentSubmissionPDF(BaseModel):
    id: str
    student_exp_id: str
    experiment_id: str
    student_id: str
    filename: str
    file_path: str
    content_type: str
    size: int
    created_at: datetime
    viewed: bool = False
    viewed_at: Optional[datetime] = None
    viewed_by: Optional[str] = None
    reviewed: bool = False
    reviewed_at: Optional[datetime] = None
    reviewed_by: Optional[str] = None
    annotations: List[PDFAnnotation] = Field(default_factory=list)


class PDFAnnotationCreateRequest(BaseModel):
    teacher_username: str
    content: str


class ResourceFile(BaseModel):
    id: str
    filename: str
    file_path: str
    file_type: str
    content_type: str
    size: int
    created_at: datetime
    created_by: str


class ResourceQuotaUpdateRequest(BaseModel):
    admin_username: str
    cpu_limit: float = Field(..., gt=0, le=128)
    memory_limit: str = Field(..., min_length=1, max_length=32)
    storage_limit: str = Field(..., min_length=1, max_length=32)
    note: Optional[str] = ""


class ResourceBudgetUpdateRequest(BaseModel):
    admin_username: str
    max_total_cpu: float = Field(..., gt=0, le=1024)
    max_total_memory: str = Field(..., min_length=1, max_length=32)
    max_total_storage: str = Field(..., min_length=1, max_length=32)
    enforce_budget: bool = False


class OperationLogEntry(BaseModel):
    id: str
    operator: str
    action: str
    target: str
    detail: str = ""
    success: bool = True
    created_at: datetime


# ==================== 模拟数据库 ====================
# 生产环境应使用 PostgreSQL + SQLAlchemy
def _hash_password(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def _default_password_hash() -> str:
    return _hash_password(DEFAULT_PASSWORD)


def _get_account_password_hash(username: str) -> str:
    normalized = _normalize_text(username)
    if not normalized:
        return _default_password_hash()

    saved_hash = _normalize_text(teacher_account_password_hashes_db.get(normalized)).lower()
    if saved_hash and PASSWORD_HASH_PATTERN.fullmatch(saved_hash):
        return saved_hash
    return _default_password_hash()


def _verify_account_password(username: str, password: str) -> bool:
    return _hash_password(password or "") == _get_account_password_hash(username)


def _normalize_security_question(question: str) -> str:
    return _normalize_text(question)[:120]


def _normalize_security_answer(answer: str) -> str:
    return _normalize_text(answer).lower()[:200]


def _hash_security_answer(answer: str) -> str:
    normalized = _normalize_security_answer(answer)
    return _hash_password(normalized)


def _verify_security_answer(stored_hash: str, provided_answer: str) -> bool:
    normalized_hash = _normalize_text(stored_hash).lower()
    if not PASSWORD_HASH_PATTERN.fullmatch(normalized_hash):
        return False
    return normalized_hash == _hash_security_answer(provided_answer)


def _normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()
def _normalize_admission_year(value) -> str:
    raw = _normalize_text(value)
    if not raw:
        return ""
    digits = "".join(ch for ch in raw if ch.isdigit())
    if len(digits) == 4 and digits.startswith("20"):
        return digits
    if len(digits) == 2:
        return f"20{digits}"
    return ""


def _all_teacher_accounts() -> List[str]:
    merged = set()
    for username in TEACHER_ACCOUNTS:
        normalized = _normalize_text(username)
        if normalized:
            merged.add(normalized)
    session_maker = _get_pg_sync_session_maker()
    if session_maker is not None:
        try:
            with session_maker() as db:
                rows = (
                    db.execute(select(UserORM).where(UserORM.role == "teacher"))
                    .scalars()
                    .all()
                )
                for row in rows:
                    normalized = _normalize_text(row.username)
                    if normalized:
                        merged.add(normalized)
        except Exception:
            pass
    return sorted(merged)


def _infer_user_role(username: str) -> str:
    normalized = _normalize_text(username)
    if is_admin(normalized):
        return "admin"
    if is_teacher(normalized):
        return "teacher"
    return "student"


def _is_student_progress_record(student_id: str) -> bool:
    """Only keep student-origin records in teacher-facing progress/review views."""
    normalized = _normalize_text(student_id)
    if not normalized:
        return False
    if is_admin(normalized) or is_teacher(normalized):
        return False
    return True


def _ensure_admin(admin_username: str):
    normalized = _normalize_text(admin_username)
    if not is_admin(normalized):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员账号")


_SIZE_LIMIT_PATTERN = re.compile(r"^\s*(\d+(?:\.\d+)?)\s*([kmgt]?b?)?\s*$", re.IGNORECASE)
_SIZE_FACTORS = {
    "B": 1,
    "K": 1024,
    "M": 1024 ** 2,
    "G": 1024 ** 3,
    "T": 1024 ** 4,
}


def _default_size_unit(default_value: str) -> str:
    match = _SIZE_LIMIT_PATTERN.match(_normalize_text(default_value))
    if not match:
        return "B"
    unit_raw = (match.group(2) or "").upper()
    if unit_raw in {"K", "KB"}:
        return "K"
    if unit_raw in {"M", "MB"}:
        return "M"
    if unit_raw in {"G", "GB"}:
        return "G"
    if unit_raw in {"T", "TB"}:
        return "T"
    return "B"


def _normalize_size_limit(value, default_value: str) -> str:
    raw = _normalize_text(value)
    if not raw:
        return default_value

    match = _SIZE_LIMIT_PATTERN.match(raw)
    if not match:
        raise HTTPException(status_code=400, detail=f"资源大小格式无效: {raw}")

    number = float(match.group(1))
    if number <= 0:
        raise HTTPException(status_code=400, detail=f"资源大小必须大于 0: {raw}")

    default_unit = _default_size_unit(default_value)
    unit_raw = (match.group(2) or "").upper()
    if unit_raw == "":
        unit = default_unit
    elif unit_raw == "B":
        # Backward compatibility: older UI values like "8" were normalized into "8B".
        unit = default_unit if default_unit != "B" else "B"
    elif unit_raw in {"K", "KB"}:
        unit = "K"
    elif unit_raw in {"M", "MB"}:
        unit = "M"
    elif unit_raw in {"G", "GB"}:
        unit = "G"
    elif unit_raw in {"T", "TB"}:
        unit = "T"
    else:
        raise HTTPException(status_code=400, detail=f"资源大小单位无效: {raw}")

    if number.is_integer():
        number_text = str(int(number))
    else:
        number_text = str(round(number, 3)).rstrip("0").rstrip(".")
    if unit == "B":
        return number_text
    return f"{number_text}{unit}"


def _size_to_bytes(value: str) -> int:
    normalized = _normalize_size_limit(value, "0B")
    match = _SIZE_LIMIT_PATTERN.match(normalized)
    if not match:
        return 0
    number = float(match.group(1))
    unit_raw = (match.group(2) or "B").upper()
    if unit_raw in {"", "B"}:
        unit = "B"
    elif unit_raw in {"K", "KB"}:
        unit = "K"
    elif unit_raw in {"M", "MB"}:
        unit = "M"
    elif unit_raw in {"G", "GB"}:
        unit = "G"
    elif unit_raw in {"T", "TB"}:
        unit = "T"
    else:
        unit = "B"
    return int(number * _SIZE_FACTORS[unit])


def _default_resource_policy_payload() -> dict:
    now_iso = datetime.now().isoformat()
    payload = {
        "defaults": deepcopy(DEFAULT_RESOURCE_ROLE_LIMITS),
        "budget": {
            **deepcopy(DEFAULT_SERVER_RESOURCE_BUDGET),
            "updated_by": "system",
            "updated_at": now_iso,
        },
        "overrides": {},
    }
    return payload


def _normalize_resource_quota(raw: Optional[dict], role: str) -> dict:
    role_key = role if role in DEFAULT_RESOURCE_ROLE_LIMITS else "student"
    default_quota = DEFAULT_RESOURCE_ROLE_LIMITS[role_key]
    source = raw or {}

    cpu_raw = source.get("cpu_limit", default_quota["cpu_limit"])
    try:
        cpu_limit = float(cpu_raw)
    except (TypeError, ValueError):
        cpu_limit = float(default_quota["cpu_limit"])
    cpu_limit = round(max(0.1, min(cpu_limit, 128.0)), 3)

    memory_limit = _normalize_size_limit(source.get("memory_limit", default_quota["memory_limit"]), default_quota["memory_limit"])
    storage_limit = _normalize_size_limit(source.get("storage_limit", default_quota["storage_limit"]), default_quota["storage_limit"])
    return {
        "cpu_limit": cpu_limit,
        "memory_limit": memory_limit,
        "storage_limit": storage_limit,
    }


def _normalize_resource_budget(raw: Optional[dict]) -> dict:
    source = raw or {}
    default_budget = DEFAULT_SERVER_RESOURCE_BUDGET
    try:
        max_total_cpu = float(source.get("max_total_cpu", default_budget["max_total_cpu"]))
    except (TypeError, ValueError):
        max_total_cpu = float(default_budget["max_total_cpu"])
    max_total_cpu = round(max(0.1, min(max_total_cpu, 1024.0)), 3)

    max_total_memory = _normalize_size_limit(
        source.get("max_total_memory", default_budget["max_total_memory"]),
        default_budget["max_total_memory"],
    )
    max_total_storage = _normalize_size_limit(
        source.get("max_total_storage", default_budget["max_total_storage"]),
        default_budget["max_total_storage"],
    )
    enforce_budget = bool(source.get("enforce_budget", default_budget["enforce_budget"]))
    updated_by = _normalize_text(source.get("updated_by")) or "system"
    updated_at = _normalize_text(source.get("updated_at")) or datetime.now().isoformat()
    return {
        "max_total_cpu": max_total_cpu,
        "max_total_memory": max_total_memory,
        "max_total_storage": max_total_storage,
        "enforce_budget": enforce_budget,
        "updated_by": updated_by,
        "updated_at": updated_at,
    }


def _save_resource_policy():
    _run_pg_write("resource_policy", lambda db: _upsert_kv(db, "resource_policy", deepcopy(resource_policy_db)))


def _load_resource_policy():
    return


def _operation_log_to_dict(record: OperationLogEntry) -> dict:
    payload = record.dict()
    payload["created_at"] = record.created_at.isoformat()
    return payload


def _save_operation_logs():
    payload = {"items": [_operation_log_to_dict(item) for item in operation_logs_db]}
    def _persist(db: Session):
        target_ids = set()
        for item in operation_logs_db:
            target_ids.add(item.id)
            db.merge(
                OperationLogORM(
                    id=item.id,
                    operator=_normalize_text(item.operator) or "system",
                    action=_normalize_text(item.action) or "unknown",
                    target=_normalize_text(item.target),
                    detail=_normalize_text(item.detail),
                    success=bool(item.success),
                    created_at=item.created_at or datetime.now(),
                )
            )

        existing = db.execute(select(OperationLogORM)).scalars().all()
        for row in existing:
            if row.id not in target_ids:
                db.delete(row)

    _run_pg_write("operation_logs", _persist)


def _load_operation_logs():
    return


def _append_operation_log(operator: str, action: str, target: str, detail: str = "", success: bool = True):
    entry = OperationLogEntry(
        id=str(uuid.uuid4()),
        operator=_normalize_text(operator) or "unknown",
        action=_normalize_text(action) or "unknown",
        target=_normalize_text(target) or "-",
        detail=_normalize_text(detail)[:800],
        success=bool(success),
        created_at=datetime.now(),
    )
    operation_logs_db.append(entry)
    if len(operation_logs_db) > MAX_OPERATION_LOG_ITEMS:
        del operation_logs_db[: len(operation_logs_db) - MAX_OPERATION_LOG_ITEMS]
    _save_operation_logs()


def _managed_users() -> List[dict]:
    users: List[dict] = []
    seen = set()

    for username in ADMIN_ACCOUNTS:
        normalized = _normalize_text(username)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        users.append({
            "username": normalized,
            "role": "admin",
            "real_name": normalized,
            "student_id": "",
            "class_name": "",
            "organization": "",
        })

    teacher_real_names: Dict[str, str] = {}
    student_rows: List[UserORM] = []
    session_maker = _get_pg_sync_session_maker()
    if session_maker is not None:
        try:
            with session_maker() as db:
                teacher_rows = (
                    db.execute(select(UserORM).where(UserORM.role == "teacher"))
                    .scalars()
                    .all()
                )
                for row in teacher_rows:
                    teacher_real_names[_normalize_text(row.username)] = _normalize_text(row.real_name)
                student_rows = list(
                    db.execute(select(UserORM).where(UserORM.role == "student"))
                    .scalars()
                    .all()
                )
        except Exception:
            teacher_real_names = {}
            student_rows = []

    for username in _all_teacher_accounts():
        normalized = _normalize_text(username)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        users.append({
            "username": normalized,
            "role": "teacher",
            "real_name": teacher_real_names.get(normalized, "") or normalized,
            "student_id": "",
            "class_name": "",
            "organization": "",
        })

    for student in student_rows:
        normalized = _normalize_text(student.username or student.student_id)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        users.append({
            "username": normalized,
            "role": "student",
            "real_name": _normalize_text(student.real_name) or normalized,
            "student_id": _normalize_text(student.student_id),
            "class_name": _normalize_text(student.class_name),
            "organization": _normalize_text(student.organization),
        })

    role_order = {"admin": 0, "teacher": 1, "student": 2}
    users.sort(key=lambda item: (role_order.get(item["role"], 9), item["username"]))
    return users


def _get_role_default_quota(role: str) -> dict:
    defaults = resource_policy_db.get("defaults", {})
    role_key = role if role in DEFAULT_RESOURCE_ROLE_LIMITS else "student"
    return _normalize_resource_quota(defaults.get(role_key), role_key)


def _get_effective_user_quota(username: str, role: str, overrides: Optional[dict] = None) -> Tuple[dict, str, dict]:
    normalized_username = _normalize_text(username)
    default_quota = _get_role_default_quota(role)
    override_source = overrides if isinstance(overrides, dict) else resource_policy_db.get("overrides", {})
    override = override_source.get(normalized_username)
    if isinstance(override, dict):
        quota = _normalize_resource_quota(override, role)
        meta = {
            "updated_by": _normalize_text(override.get("updated_by")) or "unknown",
            "updated_at": _normalize_text(override.get("updated_at")) or "",
            "note": _normalize_text(override.get("note")),
        }
        return quota, "custom", meta
    return default_quota, "default", {"updated_by": "system", "updated_at": "", "note": ""}
def _collect_resource_control_users(overrides: Optional[dict] = None) -> List[dict]:
    users = _managed_users()
    hub_map = _hub_user_state_map()
    rows = []
    for item in users:
        username = item["username"]
        role = item["role"]
        quota, source, meta = _get_effective_user_quota(username, role, overrides=overrides)
        hub_state = _extract_server_state(hub_map.get(username))
        rows.append({
            **item,
            "quota": quota,
            "quota_source": source,
            "quota_updated_by": meta.get("updated_by", ""),
            "quota_updated_at": meta.get("updated_at", ""),
            "quota_note": meta.get("note", ""),
            **hub_state,
        })
    return rows


def _resource_assignment_summary(rows: List[dict], budget: dict) -> dict:
    assigned_cpu = 0.0
    assigned_memory = 0
    assigned_storage = 0
    active_cpu = 0.0
    active_memory = 0
    active_storage = 0
    running_servers = 0

    for item in rows:
        quota = item.get("quota", {})
        cpu = float(quota.get("cpu_limit", 0.0) or 0.0)
        memory = _size_to_bytes(str(quota.get("memory_limit", "0B")))
        storage = _size_to_bytes(str(quota.get("storage_limit", "0B")))
        assigned_cpu += cpu
        assigned_memory += memory
        assigned_storage += storage
        if item.get("server_running"):
            running_servers += 1
            active_cpu += cpu
            active_memory += memory
            active_storage += storage

    budget_cpu = float(budget.get("max_total_cpu", DEFAULT_SERVER_RESOURCE_BUDGET["max_total_cpu"]))
    budget_memory = _size_to_bytes(str(budget.get("max_total_memory", DEFAULT_SERVER_RESOURCE_BUDGET["max_total_memory"])))
    budget_storage = _size_to_bytes(str(budget.get("max_total_storage", DEFAULT_SERVER_RESOURCE_BUDGET["max_total_storage"])))

    return {
        "total_users": len(rows),
        "teachers": len([item for item in rows if item["role"] == "teacher"]),
        "students": len([item for item in rows if item["role"] == "student"]),
        "admins": len([item for item in rows if item["role"] == "admin"]),
        "running_servers": running_servers,
        "assigned_cpu": round(assigned_cpu, 3),
        "assigned_memory_bytes": assigned_memory,
        "assigned_storage_bytes": assigned_storage,
        "active_cpu": round(active_cpu, 3),
        "active_memory_bytes": active_memory,
        "active_storage_bytes": active_storage,
        "budget_cpu": budget_cpu,
        "budget_memory_bytes": budget_memory,
        "budget_storage_bytes": budget_storage,
        "assigned_cpu_ratio": round((assigned_cpu / budget_cpu) if budget_cpu > 0 else 0.0, 4),
        "assigned_memory_ratio": round((assigned_memory / budget_memory) if budget_memory > 0 else 0.0, 4),
        "assigned_storage_ratio": round((assigned_storage / budget_storage) if budget_storage > 0 else 0.0, 4),
    }


def _validate_budget(summary: dict, budget: dict):
    if not budget.get("enforce_budget"):
        return
    if summary["assigned_cpu"] > summary["budget_cpu"] + 1e-9:
        raise HTTPException(status_code=409, detail="分配失败：CPU总配额超出服务器预算")
    if summary["assigned_memory_bytes"] > summary["budget_memory_bytes"]:
        raise HTTPException(status_code=409, detail="分配失败：内存总配额超出服务器预算")
    if summary["assigned_storage_bytes"] > summary["budget_storage_bytes"]:
        raise HTTPException(status_code=409, detail="分配失败：存储总配额超出服务器预算")


def _infer_admission_year(student_id: str) -> str:
    normalized_student_id = _normalize_text(student_id)
    if len(normalized_student_id) >= 2 and normalized_student_id[:2].isdigit():
        return f"20{normalized_student_id[:2]}"
    return ""


def _format_admission_year_label(admission_year: str) -> str:
    normalized = _normalize_admission_year(admission_year)
    return f"{normalized}级" if normalized else ""


def _build_class_name(admission_year: str, major_name: str, class_name: str) -> str:
    normalized_year = _normalize_admission_year(admission_year)
    normalized_major = _normalize_text(major_name)
    normalized_class = _normalize_text(class_name)
    if not (normalized_year and normalized_major and normalized_class):
        return ""
    return f"{normalized_year}级{normalized_major}{normalized_class}"


def _ensure_teacher(teacher_username: str):
    normalized_teacher = _normalize_text(teacher_username)
    if not (is_teacher(normalized_teacher) or is_admin(normalized_teacher)):
        raise HTTPException(status_code=403, detail="权限不足")


def _is_admin_user(username: str) -> bool:
    return is_admin(_normalize_text(username))


def _list_accessible_classes(teacher_username: str) -> List[ClassRecord]:
    normalized_teacher = _normalize_text(teacher_username)
    session_maker = _get_pg_sync_session_maker()
    if session_maker is None:
        return []

    try:
        with session_maker() as db:
            rows = db.execute(select(ClassroomORM)).scalars().all()
    except Exception:
        return []

    records = [
        ClassRecord(
            id=row.id,
            name=row.name,
            created_by=row.created_by,
            created_at=row.created_at,
        )
        for row in rows
    ]
    if _is_admin_user(normalized_teacher):
        return records
    return [item for item in records if _normalize_text(item.created_by) == normalized_teacher]


def _student_owner_username(record: StudentRecord) -> str:
    normalized_owner = _normalize_text(record.created_by)
    if normalized_owner:
        return normalized_owner

    session_maker = _get_pg_sync_session_maker()
    if session_maker is None:
        return ""
    try:
        with session_maker() as db:
            rows = (
                db.execute(select(ClassroomORM).where(ClassroomORM.name == record.class_name))
                .scalars()
                .all()
            )
    except Exception:
        return ""

    matched_class_owners = {
        _normalize_text(item.created_by)
        for item in rows
        if _normalize_text(item.created_by)
    }
    if len(matched_class_owners) == 1:
        return next(iter(matched_class_owners))
    return ""


def _student_visible_to_teacher(record: StudentRecord, teacher_username: str) -> bool:
    normalized_teacher = _normalize_text(teacher_username)
    if _is_admin_user(normalized_teacher):
        return True
    return _student_owner_username(record) == normalized_teacher


def _ensure_student(student_id: str):
    normalized_student_id = _normalize_text(student_id)
    if not normalized_student_id:
        raise HTTPException(status_code=404, detail="学生不存在")

    session_maker = _get_pg_sync_session_maker()
    if session_maker is None:
        raise HTTPException(status_code=404, detail="学生不存在")

    try:
        with session_maker() as db:
            row = (
                db.execute(
                    select(UserORM).where(
                        UserORM.role == "student",
                        (UserORM.student_id == normalized_student_id) | (UserORM.username == normalized_student_id),
                    )
                )
                .scalars()
                .first()
            )
    except Exception:
        row = None

    if row is None:
        raise HTTPException(status_code=404, detail="学生不存在")


def _normalize_publish_scope(value: Optional[str]) -> PublishScope:
    if isinstance(value, PublishScope):
        raw = value.value
    else:
        raw = _normalize_text(value).lower()
    if raw.startswith("publishscope."):
        raw = raw.split(".", 1)[1]
    if raw == PublishScope.CLASS.value:
        return PublishScope.CLASS
    if raw == PublishScope.STUDENT.value:
        return PublishScope.STUDENT
    return PublishScope.ALL


def _normalize_experiment_publish_targets(record: Experiment):
    record.publish_scope = _normalize_publish_scope(getattr(record, "publish_scope", PublishScope.ALL.value))

    normalized_classes: List[str] = []
    class_seen: Set[str] = set()
    for item in list(getattr(record, "target_class_names", []) or []):
        normalized = _normalize_text(item)
        key = normalized.lower()
        if not normalized or key in class_seen:
            continue
        class_seen.add(key)
        normalized_classes.append(normalized)
    record.target_class_names = normalized_classes

    normalized_students: List[str] = []
    student_seen: Set[str] = set()
    for item in list(getattr(record, "target_student_ids", []) or []):
        normalized = _normalize_text(item)
        key = normalized.lower()
        if not normalized or key in student_seen:
            continue
        student_seen.add(key)
        normalized_students.append(normalized)
    record.target_student_ids = normalized_students

    if record.publish_scope == PublishScope.ALL:
        record.target_class_names = []
        record.target_student_ids = []
    elif record.publish_scope == PublishScope.CLASS:
        record.target_student_ids = []
    elif record.publish_scope == PublishScope.STUDENT:
        record.target_class_names = []


def _validate_experiment_publish_targets(record: Experiment):
    if not record.published:
        return

    if record.publish_scope == PublishScope.CLASS and not record.target_class_names:
        raise HTTPException(status_code=400, detail="发布范围为班级时，至少选择一个班级")

    if record.publish_scope == PublishScope.STUDENT and not record.target_student_ids:
        raise HTTPException(status_code=400, detail="发布范围为学生时，至少选择一个学生")


def _is_experiment_visible_to_student(record: Experiment, student: StudentRecord) -> bool:
    if not record.published:
        return False

    _normalize_experiment_publish_targets(record)
    if record.publish_scope == PublishScope.ALL:
        return True

    if record.publish_scope == PublishScope.CLASS:
        normalized_targets = {_normalize_text(name) for name in (record.target_class_names or []) if _normalize_text(name)}
        return _normalize_text(student.class_name) in normalized_targets

    if record.publish_scope == PublishScope.STUDENT:
        normalized_targets = {_normalize_text(item) for item in (record.target_student_ids or []) if _normalize_text(item)}
        return _normalize_text(student.student_id) in normalized_targets

    return False


def _is_known_user(username: str) -> bool:
    normalized = _normalize_text(username)
    if not normalized:
        return False
    if is_teacher(normalized) or is_admin(normalized):
        return True

    session_maker = _get_pg_sync_session_maker()
    if session_maker is None:
        return False

    try:
        with session_maker() as db:
            row = (
                db.execute(
                    select(UserORM).where(
                        UserORM.role == "student",
                        (UserORM.student_id == normalized) | (UserORM.username == normalized),
                    )
                )
                .scalars()
                .first()
            )
            return row is not None
    except Exception:
        return False


def _normalize_ai_shared_config(raw: Optional[dict]) -> dict:
    payload = raw or {}
    chat_model = _normalize_text(payload.get("chat_model")) or DEFAULT_AI_SHARED_CONFIG["chat_model"]
    reasoner_model = _normalize_text(payload.get("reasoner_model")) or DEFAULT_AI_SHARED_CONFIG["reasoner_model"]
    base_url = _normalize_text(payload.get("base_url")) or DEFAULT_AI_SHARED_CONFIG["base_url"]
    system_prompt = _normalize_text(payload.get("system_prompt")) or DEFAULT_AI_SHARED_CONFIG["system_prompt"]
    api_key = _normalize_text(payload.get("api_key"))
    tavily_api_key = _normalize_text(payload.get("tavily_api_key"))

    return {
        "api_key": api_key[:512],
        "tavily_api_key": tavily_api_key[:512],
        "chat_model": chat_model[:120],
        "reasoner_model": reasoner_model[:120],
        "base_url": base_url[:500].rstrip("/") or DEFAULT_AI_SHARED_CONFIG["base_url"],
        "system_prompt": system_prompt[:4000],
    }


def _save_ai_shared_config():
    _run_pg_write("ai_shared_config", lambda db: _upsert_kv(db, "ai_shared_config", deepcopy(ai_shared_config_db)))


def _load_ai_shared_config():
    return


def _normalize_chat_history_message(raw: Optional[dict]) -> Optional[Dict[str, str]]:
    if not isinstance(raw, dict):
        return None

    role = _normalize_text(raw.get("role")).lower()
    if role not in {"system", "user", "assistant"}:
        return None

    content = str(raw.get("content") or "").strip()
    if not content:
        return None

    return {
        "role": role,
        "content": content[:AI_CHAT_HISTORY_MAX_MESSAGE_CHARS],
    }


def _normalize_chat_history_items(raw_items) -> List[Dict[str, str]]:
    output: List[Dict[str, str]] = []
    for item in raw_items if isinstance(raw_items, list) else []:
        normalized = _normalize_chat_history_message(item)
        if normalized:
            output.append(normalized)
    return output[-AI_CHAT_HISTORY_MAX_MESSAGES:]


def _save_ai_chat_history():
    payload = {}
    for username, items in ai_chat_history_db.items():
        normalized_username = _normalize_text(username)
        if not normalized_username:
            continue
        payload[normalized_username] = _normalize_chat_history_items(items)

    _run_pg_write("ai_chat_history", lambda db: _upsert_kv(db, "ai_chat_history", payload))


def _load_ai_chat_history():
    return


def _get_ai_chat_history(username: str) -> List[Dict[str, str]]:
    normalized_username = _normalize_text(username)
    if not normalized_username:
        return []
    return deepcopy(ai_chat_history_db.get(normalized_username, []))


def _set_ai_chat_history(username: str, raw_items) -> List[Dict[str, str]]:
    normalized_username = _normalize_text(username)
    if not normalized_username:
        return []
    normalized_items = _normalize_chat_history_items(raw_items)
    if normalized_items:
        ai_chat_history_db[normalized_username] = normalized_items
    else:
        ai_chat_history_db.pop(normalized_username, None)
    _save_ai_chat_history()
    return deepcopy(normalized_items)


def _trim_ai_history_for_context(raw_items) -> List[Dict[str, str]]:
    normalized_items = _normalize_chat_history_items(raw_items)
    if len(normalized_items) > AI_CONTEXT_MAX_HISTORY_MESSAGES:
        normalized_items = normalized_items[-AI_CONTEXT_MAX_HISTORY_MESSAGES:]

    total_chars = 0
    selected: List[Dict[str, str]] = []
    for item in reversed(normalized_items):
        content = item.get("content", "")
        estimated_chars = len(content) + 16
        if selected and (total_chars + estimated_chars > AI_CONTEXT_MAX_TOTAL_CHARS):
            break
        selected.append({"role": item.get("role", "user"), "content": content})
        total_chars += estimated_chars

    if not selected and normalized_items:
        last_item = normalized_items[-1]
        selected.append({
            "role": last_item.get("role", "user"),
            "content": str(last_item.get("content") or "")[:AI_CONTEXT_MAX_TOTAL_CHARS],
        })

    return list(reversed(selected))


def _class_to_dict(record: ClassRecord) -> dict:
    payload = record.dict()
    payload["created_at"] = record.created_at.isoformat()
    return payload


def _teacher_to_dict(record: TeacherRecord) -> dict:
    payload = record.dict()
    payload["created_at"] = record.created_at.isoformat()
    return payload


def _student_to_dict(record: StudentRecord) -> dict:
    payload = record.dict()
    payload["created_at"] = record.created_at.isoformat()
    payload["updated_at"] = record.updated_at.isoformat()
    return payload


def _save_user_registry():
    default_hash = _default_password_hash()
    account_password_hashes = {}
    for account, password_hash in teacher_account_password_hashes_db.items():
        normalized_account = _normalize_text(account)
        normalized_hash = _normalize_text(password_hash).lower()
        if not normalized_account:
            continue
        if not (is_teacher(normalized_account) or is_admin(normalized_account)):
            continue
        if not PASSWORD_HASH_PATTERN.fullmatch(normalized_hash):
            continue
        if normalized_hash == default_hash:
            continue
        account_password_hashes[normalized_account] = normalized_hash

    account_security_questions = {}
    for account, payload in account_security_questions_db.items():
        normalized_account = _normalize_text(account)
        if not normalized_account:
            continue
        if not (is_teacher(normalized_account) or is_admin(normalized_account)):
            continue

        raw_question = payload or {}
        normalized_question = _normalize_security_question(raw_question.get("question") or "")
        normalized_answer_hash = _normalize_text(raw_question.get("answer_hash") or "").lower()
        if not normalized_question:
            continue
        if not PASSWORD_HASH_PATTERN.fullmatch(normalized_answer_hash):
            continue

        account_security_questions[normalized_account] = {
            "question": normalized_question,
            "answer_hash": normalized_answer_hash,
        }

    payload = {
        "classes": [_class_to_dict(item) for item in classes_db.values()],
        "teachers": [_teacher_to_dict(item) for item in teachers_db.values()],
        "students": [_student_to_dict(item) for item in students_db.values()],
        "account_password_hashes": account_password_hashes,
        "account_security_questions": account_security_questions,
    }
    def _persist(db: Session):
        existing_users_by_username = {
            row.username: row
            for row in db.execute(
                select(UserORM).where(UserORM.role.in_(["teacher", "student"]))
            ).scalars().all()
        }

        class_ids = set()
        for item in payload.get("classes", []):
            class_id = _normalize_text(item.get("id"))
            if not class_id:
                continue
            class_ids.add(class_id)
            db.merge(
                ClassroomORM(
                    id=class_id,
                    name=_normalize_text(item.get("name")),
                    created_by=_normalize_text(item.get("created_by")) or "system",
                    created_at=_parse_iso_datetime(item.get("created_at")),
                )
            )

        existing_classes = db.execute(select(ClassroomORM)).scalars().all()
        for row in existing_classes:
            if row.id not in class_ids:
                db.delete(row)

        user_ids = set()
        for item in payload.get("teachers", []):
            username = _normalize_text(item.get("username"))
            if not username:
                continue
            existing = existing_users_by_username.get(username)
            user_id = existing.id if existing is not None else str(uuid.uuid5(uuid.NAMESPACE_DNS, f"teacher:{username}"))
            user_ids.add(user_id)
            db.merge(
                UserORM(
                    id=user_id,
                    username=username,
                    role="teacher",
                    real_name=_normalize_text(item.get("real_name")) or username,
                    created_by=_normalize_text(item.get("created_by")) or "system",
                    created_at=_parse_iso_datetime(item.get("created_at")),
                    updated_at=datetime.now(),
                    student_id=None,
                    class_name="",
                    admission_year="",
                    organization="",
                    phone="",
                    password_hash="",
                    security_question="",
                    security_answer_hash="",
                    is_active=True,
                    extra={},
                )
            )

        for item in payload.get("students", []):
            student_id = _normalize_text(item.get("student_id"))
            username = _normalize_text(item.get("username")) or student_id
            if not username or not student_id:
                continue
            existing = existing_users_by_username.get(username)
            user_id = existing.id if existing is not None else str(uuid.uuid5(uuid.NAMESPACE_DNS, f"student:{student_id}"))
            user_ids.add(user_id)
            db.merge(
                UserORM(
                    id=user_id,
                    username=username,
                    role="student",
                    real_name=_normalize_text(item.get("real_name")) or username,
                    student_id=student_id,
                    class_name=_normalize_text(item.get("class_name")),
                    admission_year=_normalize_text(item.get("admission_year")),
                    organization=_normalize_text(item.get("organization")),
                    phone=_normalize_text(item.get("phone")),
                    password_hash=_normalize_text(item.get("password_hash")),
                    security_question=_normalize_text(item.get("security_question")),
                    security_answer_hash=_normalize_text(item.get("security_answer_hash")),
                    created_by=_normalize_text(item.get("created_by")) or "",
                    created_at=_parse_iso_datetime(item.get("created_at")),
                    updated_at=_parse_iso_datetime(item.get("updated_at")),
                    is_active=True,
                    extra={},
                )
            )

        existing_users = db.execute(
            select(UserORM).where(UserORM.role.in_(["teacher", "student"]))
        ).scalars().all()
        for row in existing_users:
            if row.id not in user_ids:
                db.delete(row)

        _upsert_kv(db, "account_password_hashes", payload.get("account_password_hashes") or {})
        _upsert_kv(db, "account_security_questions", payload.get("account_security_questions") or {})

    _run_pg_write("user_registry", _persist)


def _load_user_registry():
    return


def _resource_to_dict(record: ResourceFile) -> dict:
    payload = record.dict()
    payload["created_at"] = record.created_at.isoformat()
    return payload


def _save_resource_registry():
    payload = {
        "items": [_resource_to_dict(item) for item in resource_files_db.values()],
    }
    def _persist(db: Session):
        target_ids = set()
        for item in payload.get("items", []):
            resource_id = _normalize_text(item.get("id"))
            if not resource_id:
                continue
            target_ids.add(resource_id)
            db.merge(
                ResourceORM(
                    id=resource_id,
                    filename=_normalize_text(item.get("filename")),
                    file_path=_normalize_text(item.get("file_path")),
                    file_type=_normalize_text(item.get("file_type")),
                    content_type=_normalize_text(item.get("content_type")),
                    size=int(item.get("size") or 0),
                    created_by=_normalize_text(item.get("created_by")),
                    created_at=_parse_iso_datetime(item.get("created_at")),
                    updated_at=datetime.now(),
                )
            )

        existing = db.execute(select(ResourceORM)).scalars().all()
        for row in existing:
            if row.id not in target_ids:
                db.delete(row)

    _run_pg_write("resource_registry", _persist)


def _load_resource_registry():
    return


def _course_to_dict(record: CourseRecord) -> dict:
    payload = record.dict()
    payload["created_at"] = record.created_at.isoformat()
    payload["updated_at"] = record.updated_at.isoformat()
    return payload


def _save_course_registry():
    payload = {
        "courses": [_course_to_dict(item) for item in courses_db.values()],
    }
    def _persist(db: Session):
        target_ids = set()
        for item in payload.get("courses", []):
            course_id = _normalize_text(item.get("id"))
            if not course_id:
                continue
            target_ids.add(course_id)
            db.merge(
                CourseORM(
                    id=course_id,
                    name=_normalize_text(item.get("name")),
                    description=_normalize_text(item.get("description")),
                    created_by=_normalize_text(item.get("created_by")),
                    created_at=_parse_iso_datetime(item.get("created_at")),
                    updated_at=_parse_iso_datetime(item.get("updated_at")),
                )
            )

        existing = db.execute(select(CourseORM)).scalars().all()
        for row in existing:
            if row.id not in target_ids:
                db.delete(row)

    _run_pg_write("course_registry", _persist)


def _load_course_registry():
    return


def _resolve_course_name(item: Experiment) -> str:
    explicit = _normalize_text(getattr(item, "course_name", ""))
    if explicit:
        return explicit

    notebook_path = _normalize_text(getattr(item, "notebook_path", ""))
    first_segment = next((seg for seg in notebook_path.split("/") if seg), "")
    if first_segment and first_segment.lower() != "course":
        return first_segment

    return "Python程序设计"


def _find_teacher_course_by_name(teacher_username: str, course_name: str) -> Optional[CourseRecord]:
    normalized_teacher = _normalize_text(teacher_username)
    normalized_name = _normalize_text(course_name).lower()
    if not normalized_teacher or not normalized_name:
        return None

    for item in courses_db.values():
        if _normalize_text(item.created_by) != normalized_teacher:
            continue
        if _normalize_text(item.name).lower() == normalized_name:
            return item
    return None


def _create_course_record(name: str, teacher_username: str, description: str = "") -> CourseRecord:
    now = datetime.now()
    record = CourseRecord(
        id=str(uuid.uuid4()),
        name=_normalize_text(name) or "未命名课程",
        description=_normalize_text(description),
        created_by=_normalize_text(teacher_username),
        created_at=now,
        updated_at=now,
    )
    courses_db[record.id] = record
    return record


def _list_course_experiments(course: CourseRecord) -> List[Experiment]:
    normalized_teacher = _normalize_text(course.created_by)
    normalized_course_id = _normalize_text(course.id)
    normalized_course_name = _normalize_text(course.name).lower()

    return [
        item
        for item in experiments_db.values()
        if _normalize_text(item.created_by) == normalized_teacher
        and (
            _normalize_text(item.course_id) == normalized_course_id
            or (
                not _normalize_text(item.course_id)
                and _resolve_course_name(item).lower() == normalized_course_name
            )
        )
    ]


def _resolve_or_create_teacher_course(
    teacher_username: str,
    course_name: str,
    requested_course_id: Optional[str] = None,
) -> Tuple[CourseRecord, bool]:
    normalized_teacher = _normalize_text(teacher_username)
    normalized_name = _normalize_text(course_name) or "Python程序设计"
    normalized_requested_id = _normalize_text(requested_course_id)

    if normalized_requested_id:
        course = courses_db.get(normalized_requested_id)
        if not course:
            raise HTTPException(status_code=404, detail="课程不存在")
        if _normalize_text(course.created_by) != normalized_teacher:
            raise HTTPException(status_code=403, detail="不能使用其他教师创建的课程")
        return course, False

    existing = _find_teacher_course_by_name(normalized_teacher, normalized_name)
    if existing:
        return existing, False

    return _create_course_record(normalized_name, normalized_teacher), True


def _sync_courses_from_experiments():
    experiments_changed = False
    courses_changed = False
    latest_activity: Dict[str, datetime] = {}

    for exp in experiments_db.values():
        teacher_username = _normalize_text(exp.created_by)
        if not teacher_username:
            continue

        course_name = _resolve_course_name(exp)
        requested_course_id = _normalize_text(exp.course_id)
        course_record: Optional[CourseRecord] = None

        if requested_course_id:
            candidate = courses_db.get(requested_course_id)
            if candidate and _normalize_text(candidate.created_by) == teacher_username:
                course_record = candidate

        if course_record is None:
            course_record = _find_teacher_course_by_name(teacher_username, course_name)

        if course_record is None:
            course_record = _create_course_record(course_name, teacher_username)
            courses_changed = True

        if exp.course_id != course_record.id:
            exp.course_id = course_record.id
            experiments_changed = True

        if _normalize_text(exp.course_name) != _normalize_text(course_record.name):
            exp.course_name = course_record.name
            experiments_changed = True

        ts = exp.created_at or datetime.now()
        previous = latest_activity.get(course_record.id)
        if previous is None or ts > previous:
            latest_activity[course_record.id] = ts

    for course_id, latest_ts in latest_activity.items():
        course = courses_db.get(course_id)
        if not course:
            continue
        if course.updated_at is None or latest_ts > course.updated_at:
            course.updated_at = latest_ts
            courses_changed = True

    if experiments_changed:
        _save_experiment_registry()
    if courses_changed:
        _save_course_registry()


def _experiment_to_dict(record: Experiment) -> dict:
    return jsonable_encoder(record)


def _save_experiment_registry():
    payload = {
        "experiments": [_experiment_to_dict(item) for item in experiments_db.values()],
    }
    def _persist(db: Session):
        target_ids = set()
        for item in payload.get("experiments", []):
            experiment_id = _normalize_text(item.get("id"))
            if not experiment_id:
                continue
            target_ids.add(experiment_id)
            db.merge(
                ExperimentORM(
                    id=experiment_id,
                    course_id=_normalize_text(item.get("course_id")) or None,
                    course_name=_normalize_text(item.get("course_name")),
                    title=_normalize_text(item.get("title")),
                    description=_normalize_text(item.get("description")),
                    difficulty=_normalize_text(item.get("difficulty")) or DifficultyLevel.BEGINNER.value,
                    tags=item.get("tags") if isinstance(item.get("tags"), list) else [],
                    notebook_path=_normalize_text(item.get("notebook_path")),
                    resources=item.get("resources") if isinstance(item.get("resources"), dict) else {},
                    deadline=_parse_iso_datetime_nullable(item.get("deadline")),
                    created_at=_parse_iso_datetime(item.get("created_at")),
                    updated_at=datetime.now(),
                    created_by=_normalize_text(item.get("created_by")) or "system",
                    published=bool(item.get("published", True)),
                    publish_scope=_normalize_text(item.get("publish_scope")) or PublishScope.ALL.value,
                    target_class_names=item.get("target_class_names") if isinstance(item.get("target_class_names"), list) else [],
                    target_student_ids=item.get("target_student_ids") if isinstance(item.get("target_student_ids"), list) else [],
                    extra={},
                )
            )

        existing = db.execute(select(ExperimentORM)).scalars().all()
        for row in existing:
            if row.id not in target_ids:
                db.delete(row)

    _run_pg_write("experiment_registry", _persist)


def _load_experiment_registry():
    return


def _get_experiment_by_notebook_path(notebook_path: str) -> Optional[Experiment]:
    needle = _normalize_text(notebook_path)
    if not needle:
        return None
    for exp in experiments_db.values():
        if _normalize_text(exp.notebook_path) == needle:
            return exp
    return None


def _ensure_default_experiments() -> bool:
    """Idempotently seed default experiments when missing."""
    from datetime import timedelta

    seeds = [
        {
            "title": "Python 基础语法练习",
            "description": "本实验旨在帮助你熟悉 Python 的基本语法，包括变量、数据类型、控制流等。",
            "difficulty": "初级",
            "tags": ["Python", "基础", "语法"],
            "notebook_path": "course/python-basics.ipynb",
            "resources": {"cpu": 1.0, "memory": "1G", "storage": "512M"},
            "deadline": datetime.now() + timedelta(days=7),
            "created_by": "teacher_001",
            "published": True,
        },
        {
            "title": "Pandas 数据分析入门",
            "description": "学习使用 Pandas 库进行基本的数据处理和分析操作，包括 DataFrame 的创建、索引、过滤等。",
            "difficulty": "中级",
            "tags": ["Data Science", "Pandas", "数据分析"],
            "notebook_path": "course/pandas-intro.ipynb",
            "resources": {"cpu": 1.0, "memory": "2G", "storage": "1G"},
            "deadline": datetime.now() + timedelta(days=14),
            "created_by": "teacher_001",
            "published": True,
        },
        {
            "title": "机器学习模型训练实战",
            "description": "使用 Scikit-learn 构建一个简单的分类模型，并在真实数据集上进行训练和评估。",
            "difficulty": "高级",
            "tags": ["Machine Learning", "Scikit-learn", "AI"],
            "notebook_path": "course/ml-training.ipynb",
            "resources": {"cpu": 2.0, "memory": "4G", "storage": "2G"},
            "deadline": datetime.now() + timedelta(days=21),
            "created_by": "teacher_001",
            "published": True,
        },
        # --- Extra lab series used by the repo's bundled notebooks ---
        {
            "title": "实验四：NumPy 数组基础与运算",
            "description": "掌握 NumPy ndarray、索引切片、广播与基础运算，为后续数据分析与可视化打基础。",
            "difficulty": "初级",
            "tags": ["Data Science", "NumPy", "实验四"],
            "notebook_path": "course/numpy-lab4.ipynb",
            "resources": {"cpu": 1.0, "memory": "2G", "storage": "1G"},
            "deadline": datetime.now() + timedelta(days=10),
            "created_by": "teacher_001",
            "published": True,
        },
        {
            "title": "实验五：Matplotlib 数据可视化",
            "description": "学习使用 Matplotlib 进行折线图、柱状图、散点图等常用可视化方法。",
            "difficulty": "中级",
            "tags": ["Data Science", "Matplotlib", "实验五"],
            "notebook_path": "course/matplotlib-lab5.ipynb",
            "resources": {"cpu": 1.0, "memory": "2G", "storage": "1G"},
            "deadline": datetime.now() + timedelta(days=12),
            "created_by": "teacher_001",
            "published": True,
        },
        {
            "title": "实验七：Pandas 数据处理与分析",
            "description": "掌握 DataFrame 基础操作、数据清洗、分组聚合与可视化分析的常用流程。",
            "difficulty": "中级",
            "tags": ["Data Science", "Pandas", "实验七"],
            "notebook_path": "course/pandas-lab7.ipynb",
            "resources": {"cpu": 1.0, "memory": "2G", "storage": "1G"},
            "deadline": datetime.now() + timedelta(days=14),
            "created_by": "teacher_001",
            "published": True,
        },
        {
            "title": "综合实验：自动驾驶视觉入门",
            "description": "综合运用 Python、数据处理与计算机视觉基础，完成一个小型自动驾驶视觉实验。",
            "difficulty": "高级",
            "tags": ["Machine Learning", "CV", "综合实验"],
            "notebook_path": "course/autodrive-vision-lab8.ipynb",
            "resources": {"cpu": 2.0, "memory": "4G", "storage": "2G"},
            "deadline": datetime.now() + timedelta(days=21),
            "created_by": "teacher_001",
            "published": True,
        },
    ]

    created = False
    for seed in seeds:
        notebook_path = seed.get("notebook_path") or ""
        if _get_experiment_by_notebook_path(notebook_path) is not None:
            continue

        exp = Experiment(**seed)
        exp.id = str(uuid.uuid4())
        exp.created_at = datetime.now()
        experiments_db[exp.id] = exp
        created = True
        print(f"[seed] Created experiment: {exp.title} ({exp.notebook_path})")

    if created:
        _save_experiment_registry()

    return created


def _normalize_file_type(file_type: str) -> str:
    if not file_type:
        return ""
    return file_type.lower().strip().lstrip(".")


def _resource_preview_mode(record: ResourceFile) -> str:
    file_type = _normalize_file_type(record.file_type)
    if file_type == "pdf":
        return "pdf"
    if file_type in {"xls", "xlsx"}:
        return "sheet"
    if file_type in {"md", "markdown"}:
        return "markdown"
    if file_type in {"txt", "csv", "json", "py", "log"}:
        return "text"
    if file_type == "docx":
        return "docx"
    return "unsupported"


def _read_text_preview(file_path: str) -> str:
    content = None
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            with open(file_path, "r", encoding=encoding) as file_obj:
                content = file_obj.read(TEXT_PREVIEW_CHAR_LIMIT + 1)
            break
        except UnicodeDecodeError:
            continue
    if content is None:
        raise HTTPException(status_code=400, detail="文本文件编码无法识别")
    if len(content) > TEXT_PREVIEW_CHAR_LIMIT:
        return f"{content[:TEXT_PREVIEW_CHAR_LIMIT]}\n\n...（预览内容已截断）"
    return content


def _read_docx_preview(file_path: str) -> str:
    try:
        with zipfile.ZipFile(file_path, "r") as archive:
            xml_bytes = archive.read("word/document.xml")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Word 文档解析失败: {exc}") from exc

    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError as exc:
        raise HTTPException(status_code=400, detail=f"Word 文档内容损坏: {exc}") from exc

    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    lines = []
    for paragraph in root.findall(".//w:p", namespace):
        text_parts = [node.text for node in paragraph.findall(".//w:t", namespace) if node.text]
        if text_parts:
            lines.append("".join(text_parts))

    text = "\n".join(lines)
    if len(text) > TEXT_PREVIEW_CHAR_LIMIT:
        return f"{text[:TEXT_PREVIEW_CHAR_LIMIT]}\n\n...（预览内容已截断）"
    return text


def _resource_to_payload(record: ResourceFile, route_prefix: str = "/api/admin/resources") -> dict:
    normalized_prefix = route_prefix.rstrip("/")
    preview_mode = _resource_preview_mode(record)
    return {
        "id": record.id,
        "filename": record.filename,
        "file_type": record.file_type,
        "content_type": record.content_type,
        "size": record.size,
        "created_at": record.created_at,
        "created_by": record.created_by,
        "preview_mode": preview_mode,
        "previewable": preview_mode != "unsupported",
        "preview_url": f"{normalized_prefix}/{record.id}/preview",
        "download_url": f"{normalized_prefix}/{record.id}/download",
    }


def _get_resource_or_404(resource_id: str) -> ResourceFile:
    record = resource_files_db.get(resource_id)
    if not record:
        raise HTTPException(status_code=404, detail="资源文件不存在")
    return record


def _ensure_resource_file_exists(record: ResourceFile):
    if not os.path.exists(record.file_path):
        resource_files_db.pop(record.id, None)
        _save_resource_registry()
        raise HTTPException(status_code=404, detail="资源文件不存在")


def _list_resource_records(name_filter: str = "", type_filter: str = "") -> List[ResourceFile]:
    normalized_name_filter = (name_filter or "").strip().lower()
    normalized_type_filter = _normalize_file_type(type_filter or "")

    items = []
    for record in resource_files_db.values():
        if normalized_name_filter and normalized_name_filter not in record.filename.lower():
            continue
        if normalized_type_filter and _normalize_file_type(record.file_type) != normalized_type_filter:
            continue
        if not os.path.exists(record.file_path):
            continue
        items.append(record)

    items.sort(key=lambda item: item.created_at, reverse=True)
    return items


def _is_template_header(row_values: List[str]) -> bool:
    normalized = [_normalize_text(value) for value in row_values[:len(TEMPLATE_HEADERS)]]
    return (
        normalized[:len(TEMPLATE_HEADERS)] == TEMPLATE_HEADERS
        or normalized[:len(LEGACY_TEMPLATE_HEADERS)] == LEGACY_TEMPLATE_HEADERS
    )


def _read_rows_from_csv(file_content: bytes) -> List[Tuple[int, List[str]]]:
    try:
        content = file_content.decode("utf-8-sig")
    except UnicodeDecodeError:
        content = file_content.decode("gbk")

    parsed_rows: List[Tuple[int, List[str]]] = []
    column_count = len(TEMPLATE_HEADERS)
    reader = csv.reader(io.StringIO(content))
    for row_index, row in enumerate(reader, start=1):
        values = [_normalize_text(value) for value in row[:column_count]]
        while len(values) < column_count:
            values.append("")
        if row_index == 1 and _is_template_header(values):
            continue
        if not any(values):
            continue
        parsed_rows.append((row_index, values))
    return parsed_rows


def _read_rows_from_xlsx(file_content: bytes) -> List[Tuple[int, List[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl is required for xlsx support") from exc

    parsed_rows: List[Tuple[int, List[str]]] = []
    column_count = len(TEMPLATE_HEADERS)
    workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    sheet = workbook.active
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [_normalize_text(value) for value in list(row)[:column_count]]
        while len(values) < column_count:
            values.append("")
        if row_index == 1 and _is_template_header(values):
            continue
        if not any(values):
            continue
        parsed_rows.append((row_index, values))
    workbook.close()
    return parsed_rows


def _parse_student_import_rows(filename: str, file_content: bytes) -> List[Tuple[int, List[str]]]:
    extension = os.path.splitext((filename or "").lower())[1]
    if extension == ".csv":
        return _read_rows_from_csv(file_content)
    if extension == ".xlsx":
        return _read_rows_from_xlsx(file_content)
    raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .csv 文件")


def _build_csv_template() -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(TEMPLATE_HEADERS)
    return buffer.getvalue().encode("utf-8-sig")


def _build_xlsx_template() -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl is required for xlsx support") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "students"
    sheet.append(TEMPLATE_HEADERS)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _is_class_template_header(row_values: List[str]) -> bool:
    normalized = [_normalize_text(value) for value in row_values[:len(CLASS_TEMPLATE_HEADERS)]]
    return normalized == CLASS_TEMPLATE_HEADERS


def _read_class_rows_from_csv(file_content: bytes) -> List[Tuple[int, List[str]]]:
    try:
        content = file_content.decode("utf-8-sig")
    except UnicodeDecodeError:
        content = file_content.decode("gbk")

    parsed_rows: List[Tuple[int, List[str]]] = []
    column_count = len(CLASS_TEMPLATE_HEADERS)
    reader = csv.reader(io.StringIO(content))
    for row_index, row in enumerate(reader, start=1):
        values = [_normalize_text(value) for value in row[:column_count]]
        while len(values) < column_count:
            values.append("")
        if row_index == 1 and _is_class_template_header(values):
            continue
        if not any(values):
            continue
        parsed_rows.append((row_index, values))
    return parsed_rows


def _read_class_rows_from_xlsx(file_content: bytes) -> List[Tuple[int, List[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl is required for xlsx support") from exc

    parsed_rows: List[Tuple[int, List[str]]] = []
    column_count = len(CLASS_TEMPLATE_HEADERS)
    workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    sheet = workbook.active
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [_normalize_text(value) for value in list(row)[:column_count]]
        while len(values) < column_count:
            values.append("")
        if row_index == 1 and _is_class_template_header(values):
            continue
        if not any(values):
            continue
        parsed_rows.append((row_index, values))
    workbook.close()
    return parsed_rows


def _parse_class_import_rows(filename: str, file_content: bytes) -> List[Tuple[int, List[str]]]:
    extension = os.path.splitext((filename or "").lower())[1]
    if extension == ".csv":
        return _read_class_rows_from_csv(file_content)
    if extension == ".xlsx":
        return _read_class_rows_from_xlsx(file_content)
    raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .csv 文件")


def _build_class_csv_template() -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(CLASS_TEMPLATE_HEADERS)
    return buffer.getvalue().encode("utf-8-sig")


def _build_class_xlsx_template() -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl is required for xlsx support") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "classes"
    sheet.append(CLASS_TEMPLATE_HEADERS)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _student_experiment_to_dict(record: StudentExperiment) -> dict:
    return jsonable_encoder(record)


def _submission_pdf_to_dict(record: StudentSubmissionPDF) -> dict:
    return jsonable_encoder(record)


def _save_submission_registries():
    submission_payload = {
        "items": [_student_experiment_to_dict(item) for item in student_experiments_db.values()],
    }
    submission_pdf_payload = {
        "items": [_submission_pdf_to_dict(item) for item in submission_pdfs_db.values()],
    }

    def _persist(db: Session):
        target_submission_ids = set()
        for item in submission_payload.get("items", []):
            submission_id = _normalize_text(item.get("id"))
            experiment_id = _normalize_text(item.get("experiment_id"))
            student_id = _normalize_text(item.get("student_id"))
            if not submission_id or not experiment_id or not student_id:
                continue
            target_submission_ids.add(submission_id)
            status_value = item.get("status")
            if isinstance(status_value, dict):
                status_text = _normalize_text(status_value.get("value"))
            else:
                status_text = _normalize_text(status_value)
            db.merge(
                SubmissionORM(
                    id=submission_id,
                    experiment_id=experiment_id,
                    student_id=student_id,
                    status=status_text or ExperimentStatus.NOT_STARTED.value,
                    start_time=_parse_iso_datetime_nullable(item.get("start_time")),
                    submit_time=_parse_iso_datetime_nullable(item.get("submit_time")),
                    notebook_content=_normalize_text(item.get("notebook_content")),
                    score=item.get("score"),
                    ai_feedback=_normalize_text(item.get("ai_feedback")),
                    teacher_comment=_normalize_text(item.get("teacher_comment")),
                    created_at=_parse_iso_datetime_nullable(item.get("start_time")) or datetime.now(),
                    updated_at=_parse_iso_datetime_nullable(item.get("submit_time")) or datetime.now(),
                    extra={},
                )
            )

        existing_submissions = db.execute(select(SubmissionORM)).scalars().all()
        for row in existing_submissions:
            if row.id not in target_submission_ids:
                db.delete(row)

        target_pdf_ids = set()
        for item in submission_pdf_payload.get("items", []):
            pdf_id = _normalize_text(item.get("id"))
            submission_id = _normalize_text(item.get("student_exp_id"))
            if not pdf_id or not submission_id or submission_id not in target_submission_ids:
                continue
            target_pdf_ids.add(pdf_id)
            db.merge(
                SubmissionPdfORM(
                    id=pdf_id,
                    submission_id=submission_id,
                    experiment_id=_normalize_text(item.get("experiment_id")),
                    student_id=_normalize_text(item.get("student_id")),
                    filename=_normalize_text(item.get("filename")),
                    file_path=_normalize_text(item.get("file_path")),
                    content_type=_normalize_text(item.get("content_type")),
                    size=int(item.get("size") or 0),
                    viewed=bool(item.get("viewed", False)),
                    viewed_at=_parse_iso_datetime_nullable(item.get("viewed_at")),
                    viewed_by=_normalize_text(item.get("viewed_by")),
                    reviewed=bool(item.get("reviewed", False)),
                    reviewed_at=_parse_iso_datetime_nullable(item.get("reviewed_at")),
                    reviewed_by=_normalize_text(item.get("reviewed_by")),
                    annotations=item.get("annotations") if isinstance(item.get("annotations"), list) else [],
                    created_at=_parse_iso_datetime(item.get("created_at")),
                    updated_at=datetime.now(),
                )
            )

        existing_pdfs = db.execute(select(SubmissionPdfORM)).scalars().all()
        for row in existing_pdfs:
            if row.id not in target_pdf_ids:
                db.delete(row)

    _run_pg_write("submission_registries", _persist)


def _load_submission_registry():
    return


def _load_submission_pdf_registry():
    return


def _get_submission_pdfs(student_exp_id: str) -> List[StudentSubmissionPDF]:
    items = [
        item for item in submission_pdfs_db.values()
        if item.student_exp_id == student_exp_id
    ]
    items.sort(key=lambda item: item.created_at, reverse=True)
    return items


def _pdf_status(item: StudentSubmissionPDF) -> str:
    if item.reviewed:
        return "已批阅"
    if item.viewed:
        return "已查看"
    return "未查看"


def _pdf_to_payload(item: StudentSubmissionPDF) -> dict:
    return {
        "id": item.id,
        "student_exp_id": item.student_exp_id,
        "experiment_id": item.experiment_id,
        "student_id": item.student_id,
        "filename": item.filename,
        "content_type": item.content_type,
        "size": item.size,
        "created_at": item.created_at,
        "download_url": f"/api/student-submissions/{item.id}/download",
        "viewed": item.viewed,
        "viewed_at": item.viewed_at,
        "viewed_by": item.viewed_by,
        "reviewed": item.reviewed,
        "reviewed_at": item.reviewed_at,
        "reviewed_by": item.reviewed_by,
        "review_status": _pdf_status(item),
        "annotations": [
            {
                "id": ann.id,
                "teacher_username": ann.teacher_username,
                "content": ann.content,
                "created_at": ann.created_at,
            }
            for ann in item.annotations
        ],
    }

# ==================== 初始化数据 ====================
def _list_admin_teacher_items() -> List[dict]:
    rows: List[dict] = []
    for username in _all_teacher_accounts():
        teacher = teachers_db.get(username)
        is_registry = teacher is not None
        rows.append({
            "username": username,
            "real_name": _normalize_text(getattr(teacher, "real_name", "")) or username,
            "source": "registry" if is_registry else "env",
            "created_by": _normalize_text(getattr(teacher, "created_by", "")) or ("system" if is_registry else "env"),
            "created_at": getattr(teacher, "created_at", None),
        })
    return rows











































# ---------- 实验管理 ----------

# ---------- 学生实验管理 ----------

























# ---------- 教师功能 ----------



# ---------- 教师课程管理 ----------

def _course_to_payload(course: CourseRecord) -> dict:
    experiments = sorted(
        _list_course_experiments(course),
        key=lambda item: item.created_at or datetime.min,
        reverse=True,
    )
    published_count = sum(1 for item in experiments if item.published)
    latest_experiment_at = experiments[0].created_at if experiments else None
    tags = sorted(
        {
            tag
            for item in experiments
            for tag in (item.tags or [])
            if _normalize_text(tag)
        }
    )
    return {
        "id": course.id,
        "name": course.name,
        "description": course.description or "",
        "created_by": course.created_by,
        "created_at": course.created_at,
        "updated_at": course.updated_at,
        "experiment_count": len(experiments),
        "published_count": published_count,
        "latest_experiment_at": latest_experiment_at,
        "tags": tags,
        "experiments": experiments,
    }

















# ==================== 附件管理 ====================

class Attachment(BaseModel):
    id: str
    experiment_id: str
    filename: str
    file_path: str
    content_type: str
    size: int
    created_at: datetime

def _attachment_to_dict(record: Attachment) -> dict:
    return jsonable_encoder(record)


def _save_attachment_registry():
    payload = {
        "items": [
            _attachment_to_dict(item)
            for item in attachments_db.values()
            if os.path.exists(item.file_path)
        ],
    }
    def _persist(db: Session):
        target_ids = set()
        for item in payload.get("items", []):
            attachment_id = _normalize_text(item.get("id"))
            experiment_id = _normalize_text(item.get("experiment_id"))
            if not attachment_id or not experiment_id:
                continue
            target_ids.add(attachment_id)
            db.merge(
                AttachmentORM(
                    id=attachment_id,
                    experiment_id=experiment_id,
                    filename=_normalize_text(item.get("filename")),
                    file_path=_normalize_text(item.get("file_path")),
                    content_type=_normalize_text(item.get("content_type")),
                    size=int(item.get("size") or 0),
                    created_at=_parse_iso_datetime(item.get("created_at")),
                    updated_at=datetime.now(),
                )
            )

        existing = db.execute(select(AttachmentORM)).scalars().all()
        for row in existing:
            if row.id not in target_ids:
                db.delete(row)

    _run_pg_write("attachment_registry", _persist)


def _load_attachment_registry():
    return


def _find_resource_by_filename(filename: str) -> Optional[ResourceFile]:
    needle = _normalize_text(filename)
    if not needle:
        return None
    for record in resource_files_db.values():
        if _normalize_text(record.filename) == needle:
            return record
    return None


def _find_latest_upload_path_by_suffix(filename_suffix: str) -> Optional[str]:
    """Find latest file in UPLOAD_DIR with the given suffix (case-insensitive)."""
    suffix = _normalize_text(filename_suffix).lower()
    if not suffix:
        return None

    candidates = []
    try:
        entries = os.listdir(UPLOAD_DIR)
    except OSError:
        return None

    for entry in entries:
        if entry.lower().endswith(suffix):
            full_path = os.path.join(UPLOAD_DIR, entry)
            if os.path.isfile(full_path):
                candidates.append(full_path)

    if not candidates:
        return None

    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]


def _attachment_exists(experiment_id: str, filename: str) -> bool:
    exp_id = _normalize_text(experiment_id)
    needle = _normalize_text(filename)
    if not (exp_id and needle):
        return False
    for item in attachments_db.values():
        if item.experiment_id == exp_id and _normalize_text(item.filename) == needle:
            return True
    return False


def _create_attachment_from_file(
    *,
    experiment_id: str,
    filename: str,
    source_path: str,
    content_type: str = "",
) -> Optional[Attachment]:
    exp_id = _normalize_text(experiment_id)
    fname = os.path.basename(_normalize_text(filename))
    src = _normalize_text(source_path)
    if not (exp_id and fname and src):
        return None
    if not os.path.exists(src):
        return None

    att_id = str(uuid.uuid4())
    safe_filename = fname.replace(" ", "_").replace("/", "_").replace("\\", "_")
    dest_path = os.path.join(UPLOAD_DIR, f"{att_id}_{safe_filename}")
    try:
        shutil.copyfile(src, dest_path)
    except Exception as exc:
        print(f"Failed to copy attachment source {src} -> {dest_path}: {exc}")
        return None

    media_type = content_type or mimetypes.guess_type(fname)[0] or "application/octet-stream"
    size = os.path.getsize(dest_path)
    record = Attachment(
        id=att_id,
        experiment_id=exp_id,
        filename=fname,
        file_path=dest_path,
        content_type=media_type,
        size=size,
        created_at=datetime.now(),
    )
    attachments_db[record.id] = record
    return record


def _any_attachment_exists(experiment_id: str, filenames: List[str]) -> bool:
    for name in filenames:
        if _attachment_exists(experiment_id, name):
            return True
    return False


def _ensure_default_attachments() -> bool:
    """Attach bundled guide/docs/templates to seeded experiments (idempotent)."""
    defaults = [
        {
            "notebook_path": "course/numpy-lab4.ipynb",
            "guides": ["lab4-numpy-guide.pdf", "lab4-numpy-guide.doc"],
            "templates": ["学号+姓名+实验四_202510.doc", "学号+姓名+实验四_202510.docx"],
        },
        {
            "notebook_path": "course/matplotlib-lab5.ipynb",
            "guides": ["lab5-matplotlib-guide.pdf", "lab5-matplotlib-guide.doc"],
            "templates": ["学号+姓名+实验五_202511.doc", "学号+姓名+实验五_202511.docx"],
        },
        {
            "notebook_path": "course/pandas-lab7.ipynb",
            "guides": ["lab7-pandas-guide.pdf", "lab7-pandas-guide.doc"],
            "templates": ["学号+姓名+实验七.doc", "学号+姓名+实验七.docx"],
        },
        {
            "notebook_path": "course/autodrive-vision-lab8.ipynb",
            "guides": ["lab8-autodrive-guide.pdf", "lab8-autodrive-guide.doc"],
            "templates": ["学号+姓名+综合实验.doc", "学号+姓名+综合实验.docx"],
        },
    ]

    created_any = False

    for entry in defaults:
        notebook_path = entry.get("notebook_path") or ""
        exp = _get_experiment_by_notebook_path(notebook_path)
        if exp is None:
            continue

        # Guides: attach at most one (prefer PDF, fallback to Word).
        guide_candidates = entry.get("guides", [])
        if not _any_attachment_exists(exp.id, guide_candidates):
            for guide_name in guide_candidates:
                src_path = _find_latest_upload_path_by_suffix(guide_name)
                if not src_path:
                    continue
                record = _create_attachment_from_file(
                    experiment_id=exp.id,
                    filename=guide_name,
                    source_path=src_path,
                    content_type=mimetypes.guess_type(guide_name)[0] or "",
                )
                if record:
                    created_any = True
                    print(f"[seed] Attached {guide_name} -> {exp.title}")
                    break

        # Templates: prefer .doc when present, fallback to .docx.
        template_names = entry.get("templates", [])
        if not _any_attachment_exists(exp.id, template_names):
            for template_name in template_names:
                resource = _find_resource_by_filename(template_name)
                if resource is None:
                    continue
                record = _create_attachment_from_file(
                    experiment_id=exp.id,
                    filename=resource.filename,
                    source_path=resource.file_path,
                    content_type=resource.content_type or "",
                )
                if record:
                    created_any = True
                    print(f"[seed] Attached {resource.filename} -> {exp.title}")
                    break  # Only attach one template variant (.doc preferred via ordering).

    if created_any:
        _save_attachment_registry()

    return created_any


def _cleanup_seeded_attachments() -> bool:
    """Remove duplicate seeded attachments (e.g. doc+docx templates, pdf+doc guides).

    Keep at most one guide (prefer PDF) and one template (prefer .doc, fallback .docx).
    Only touches known seeded filenames.
    """
    groups = [
        {
            "notebook_path": "course/numpy-lab4.ipynb",
            "guides": ["lab4-numpy-guide.pdf", "lab4-numpy-guide.doc"],
            "templates": ["学号+姓名+实验四_202510.doc", "学号+姓名+实验四_202510.docx"],
        },
        {
            "notebook_path": "course/matplotlib-lab5.ipynb",
            "guides": ["lab5-matplotlib-guide.pdf", "lab5-matplotlib-guide.doc"],
            "templates": ["学号+姓名+实验五_202511.doc", "学号+姓名+实验五_202511.docx"],
        },
        {
            "notebook_path": "course/pandas-lab7.ipynb",
            "guides": ["lab7-pandas-guide.pdf", "lab7-pandas-guide.doc"],
            "templates": ["学号+姓名+实验七.doc", "学号+姓名+实验七.docx"],
        },
        {
            "notebook_path": "course/autodrive-vision-lab8.ipynb",
            "guides": ["lab8-autodrive-guide.pdf", "lab8-autodrive-guide.doc"],
            "templates": ["学号+姓名+综合实验.doc", "学号+姓名+综合实验.docx"],
        },
    ]

    removed_any = False

    def _remove_by_filename(exp_id: str, filename: str):
        nonlocal removed_any
        needle = _normalize_text(filename)
        if not needle:
            return
        to_remove = [
            (att_id, item)
            for att_id, item in attachments_db.items()
            if item.experiment_id == exp_id and _normalize_text(item.filename) == needle
        ]
        for att_id, item in to_remove:
            attachments_db.pop(att_id, None)
            removed_any = True
            if item and os.path.exists(item.file_path):
                try:
                    os.remove(item.file_path)
                except OSError:
                    pass

    for group in groups:
        exp = _get_experiment_by_notebook_path(group.get("notebook_path") or "")
        if exp is None:
            continue

        # Guide: prefer PDF, otherwise keep the first existing in list order.
        guide_candidates = group.get("guides", [])
        guide_keep = None
        for name in guide_candidates:
            if _attachment_exists(exp.id, name):
                guide_keep = name
                break
        if guide_keep:
            for name in guide_candidates:
                if name != guide_keep and _attachment_exists(exp.id, name):
                    _remove_by_filename(exp.id, name)

        # Template: prefer .doc, fallback .docx.
        template_candidates = group.get("templates", [])
        template_keep = None
        for name in template_candidates:
            if _attachment_exists(exp.id, name):
                template_keep = name
                break
        if template_keep:
            for name in template_candidates:
                if name != template_keep and _attachment_exists(exp.id, name):
                    _remove_by_filename(exp.id, name)

    if removed_any:
        _save_attachment_registry()

    return removed_any


def _is_pdf_attachment(attachment: Attachment) -> bool:
    lower_filename = attachment.filename.lower()
    return attachment.content_type == "application/pdf" or lower_filename.endswith(".pdf")


def _is_word_filename(filename: str) -> bool:
    lower_filename = filename.lower()
    return lower_filename.endswith(".docx") or lower_filename.endswith(".doc")


def _find_paired_word_attachment(pdf_attachment: Attachment) -> Optional[Attachment]:
    base_name = os.path.splitext(pdf_attachment.filename)[0]
    candidates: List[Attachment] = []

    for item in attachments_db.values():
        if item.experiment_id != pdf_attachment.experiment_id:
            continue
        if os.path.splitext(item.filename)[0] != base_name:
            continue
        if not _is_word_filename(item.filename):
            continue
        if not os.path.exists(item.file_path):
            continue
        candidates.append(item)

    if not candidates:
        return None

    # Prefer docx when both doc and docx exist.
    candidates.sort(key=lambda item: 0 if item.filename.lower().endswith(".docx") else 1)
    return candidates[0]





# ==================== 资源文件管理 ====================






















# ==================== AI集成接口 ====================
